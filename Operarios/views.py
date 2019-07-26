import logging
from django.shortcuts import render, redirect, render_to_response
from django.core import serializers
from datetime import datetime
from django.http import HttpResponse, Http404, JsonResponse
from django.contrib import messages
from django.urls import reverse_lazy
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.forms.models import inlineformset_factory
from django.template import RequestContext
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.contrib.auth.models import User
from openpyxl import Workbook
from Operarios.models import Ciudad, Cliente, Nacionalidad
from Operarios.models import PuntoServicio, Operario, RelevamientoCab, RelevamientoDet, RelevamientoEsp, RelevamientoCupoHoras, RelevamientoMensualeros, PlanificacionCab, PlanificacionOpe, PlanificacionEsp, Cargo, CargoAsignado, AsigFiscalPuntoServicio, AsigJefeFiscal, AsignacionCab, AsignacionDet,EsmeEmMarcaciones, Feriados
from django.db import connection, transaction

from Operarios.forms import PuntoServicioForm, OperarioForm, RelevamientoForm, RelevamientoDetForm, RelevamientoEspForm, RelevamientoCupoHorasForm, RelevamientoMensualerosForm, PlanificacionForm, PlanificacionOpeForm, PlanificacionEspForm, AsignacionCabForm, AsignacionDetForm

import json
import datetime
import decimal
from django.core.paginator import Paginator
from django.forms.models import model_to_dict
from datetime import datetime
from django.db.models import Q
def index(request):
    return HttpResponse("Vista de Operarios")

def index_alert(request):
    messages.debug(request, ' SQL statements were executed.')
    messages.info(request, 'Mensaje Informativo.')
    messages.success(request, 'Mensaje de Exito')
    messages.warning(request, 'Mensaje de Error.')
    messages.error(request, 'Document deleted.')
    return render(request, 'base_example.html')

@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('Operarios.view_puntoservicio', raise_exception=True), name='dispatch')
class PuntosServicioList(ListView):
    #model = PuntoServicio
    context_object_name = 'PuntoServicio'
    template_name = "puntoServicio/puntoServicio_list.html"
    def get_queryset(self):
        try:
            #Traemos el cargo asignado
            ConsultaCargoUsuario = Cargo.objects.filter(cargoasignado__user=self.request.user.id).query
            logging.getLogger("error_logger").error('La consulta del cargo del usuario logueado ejecutada es: {0}'.format(ConsultaCargoUsuario))
            cargoUsuario = Cargo.objects.get(cargoasignado__user=self.request.user.id)

            if (cargoUsuario.cargo == 'Fiscal'):
                #Si es fiscal, le trae sus puntos de servicios. 'Fiscal'
                consulta = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))

                return PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id))

            elif (cargoUsuario.cargo == 'Jefe de Operaciones'):
                #Si es jefe de operaciones -> Trae todo los puntos de servicio de sus fiscales. 'Jefe de Operaciones'
                consulta = PuntoServicio.objects.filter( Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)  ).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))
                return PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id) )
                pass
            elif (cargoUsuario.cargo == 'Gerente de Operaciones'):
                #Si es Gerente de Operaciones/SubGerente -> Todos los contratos. 'Gerente de Operaciones'
                consulta = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)  ).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))
                return PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id) )
                pass

            else:
                #Else: algun error, de no tener el rol correcto.
                pass
        except Cargo.DoesNotExist:
            raise Http404("El usuario no tiene el cargo requerido para ingresar")        

@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('Operarios.view_puntoservicio', raise_exception=True), name='dispatch')
class PuntosServicioAprobados(ListView):
    #model = PuntoServicio
    context_object_name = 'PuntoServicio'
    template_name = "puntoServicio/puntoServicio_aprobado.html"
    def get_queryset(self):
        try:
            #Traemos el cargo asignado
            ConsultaCargoUsuario = Cargo.objects.filter(cargoasignado__user=self.request.user.id).query
            logging.getLogger("error_logger").error('La consulta del cargo del usuario logueado ejecutada es: {0}'.format(ConsultaCargoUsuario))
            cargoUsuario = Cargo.objects.get(cargoasignado__user=self.request.user.id)

            if (cargoUsuario.cargo == 'Fiscal'):
                #Si es fiscal, le trae sus puntos de servicios. 'Fiscal'
                consulta = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)  ).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))
                return PuntoServicio.objects.filter( Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id) )
            elif (cargoUsuario.cargo == 'Jefe de Operaciones'):
                #Si es jefe de operaciones -> Trae todo los puntos de servicio de sus fiscales. 'Jefe de Operaciones'
                consulta = PuntoServicio.objects.filter( Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))
                return PuntoServicio.objects.filter( Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id))
                pass
            elif (cargoUsuario.cargo == 'Gerente de Operaciones'):
                #Si es Gerente de Operaciones/SubGerente -> Todos los contratos. 'Gerente de Operaciones'
                consulta = PuntoServicio.objects.filter( Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id)).query
                logging.getLogger("error_logger").error('La consulta de puntos de Servicio List ejecutada es: {0}'.format(consulta))
                return PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=self.request.user.id))
                pass

            else:
                #Else: algun error, de no tener el rol correcto.
                pass
        except Cargo.DoesNotExist:
            raise Http404("El usuario no tiene el cargo requerido para ingresar")  


@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('Operarios.add_puntoservicio', raise_exception=True), name='dispatch')
class PuntoServicioCreate(SuccessMessageMixin, CreateView):
    model = PuntoServicio
    ciudades = Ciudad.objects.all()
    clientes = Cliente.objects.all()
    form_class = PuntoServicioForm
    template_name = "puntoServicio/puntoServicio_form.html"
    success_url = reverse_lazy('Operarios:puntoServicio_list')
    success_message = 'Punto de Servicio Creado correctamente'
    extra_context = {'title': 'Nuevo Punto de Servicio','ciudades':ciudades,'clientes':clientes}

@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('Operarios.change_puntoservicio', raise_exception=True), name='dispatch')
class PuntoServicioUpdateView(SuccessMessageMixin, UpdateView):
    model = PuntoServicio
    form_class = PuntoServicioForm
    template_name = "puntoServicio/puntoServicio_form.html"
    success_url = reverse_lazy('Operarios:puntoServicio_list')
    success_message = 'Punto de Servicio modificado correctamente'
    extra_context = {'title': 'Editar Punto de Servicio '}


@login_required
@permission_required('Operarios.change_puntoservicio', raise_exception=True)
def PuntosServicios_update(request, pk):
    puntoServicio = PuntoServicio.objects.get(Q(id=pk))
  
    if request.method == 'GET':
        form = PuntoServicioForm(instance=puntoServicio)
        contexto = {
            'title': 'Editar Punto de servicio',
            'form': form,
            'puntoServicio':puntoServicio 
        }
    else:
        form = PuntoServicioForm(request.POST, instance=puntoServicio)
        if form.is_valid():
            #form.save()
            print(form.cleaned_data)
            conn= connection.cursor()
            params=(pk,form.cleaned_data.get('CodPuntoServicio'),
            form.cleaned_data.get('NombrePServicio'),
            form.cleaned_data.get('DireccionContrato'),
            form.cleaned_data.get('Barrios'),
            form.cleaned_data.get('Contacto'),
            form.cleaned_data.get('MailContacto'),
            form.cleaned_data.get('TelefonoContacto'),
            form.cleaned_data.get('Coordenadas'),
            str(form.cleaned_data.get('Ciudad')),
            str(form.cleaned_data.get('Cliente')),
            form.cleaned_data.get('NumeroMarcador'),
            int(0),int(0));
            print(params)
            conn.execute('puntoServicio_trigger %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s',params)
            result = conn.fetchone()[0]
            conn.close()
            if result==0:
                messages.success(request, 'Punto de Servicio modificado correctamente.')    
            else:
                messages.success(request, 'Error al modificar Punto de Servicio.')    
        return redirect('Operarios:puntoServicio_list')

    return render(request, 'puntoServicio/puntoServicio_form.html', context=contexto) 

@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('Operarios.delete_puntoservicio', raise_exception=True), name='dispatch')
class PuntoServicioDeleteView(SuccessMessageMixin, DeleteView):
    model = PuntoServicio
    template_name = "puntoServicio/puntoServicio_delete.html"
    success_url = reverse_lazy('Operarios:puntoServicio_list')
    success_message = 'Punto de Servicio eliminado correctamente'

    def delete(self, request, *args, **kwargs):
        messages.warning(self.request, self.success_message)
        return super(PuntoServicioDeleteView, self).delete(request, *args, **kwargs)


@login_required
@permission_required('Operarios.view_puntoservicio', raise_exception=True)
def PuntoServicio_list(request):
    puntoServicio = PuntoServicio.objects.all()
    print(request.GET)
    if request.GET.get('id'):
        puntoServicio=puntoServicio.filter(id__contains=request.GET.get('id'))
    
    if request.GET.get('codigoPuntoServicio'):
        puntoServicio=puntoServicio.filter(CodPuntoServicio__contains=request.GET.get('codigoPuntoServicio'))
    
    if request.GET.get('nombrePuntoServicio'):
       puntoServicio=puntoServicio.filter(NombrePServicio__contains=request.GET.get('nombrePuntoServicio'))

    if request.GET.get('Cliente'):
       puntoServicio=puntoServicio.filter(nombre__contains=request.GET.get('Cliente'))


    if request.GET.get('clientePuntoServicio'):
        puntoServicio=puntoServicio.filter(Cliente_id=request.GET.get('clientePuntoServicio'))

    """
    Se reestructura para obtener nombre de cliente
    """
    puntos = []
    for p in puntoServicio:
        cliente = ""
        if p.Cliente:
            cliente = Cliente.objects.get(id=p.Cliente.id)

        puntos.append({
            "id":p.id,
            "codigoPuntoServicio":p.CodPuntoServicio,
            "nombrePuntoServicio":p.NombrePServicio,
            "clientePuntoServicio":cliente.id
        })
    
    # paginado=Paginator(puntoServicio.order_by('id').values("id", "CodPuntoServicio", "NombrePServicio","Cliente"),  request.GET.get('pageSize'))
    # listaPaginada=paginado.page(request.GET.get('pageIndex')).object_list
    # dataPuntosServicio=list(puntos)
    
    response={}
    response['dato']=puntos
    return HttpResponse(json.dumps(response),content_type="application/json")

 
def getClientes(request):
    clientes = Cliente.objects.all()
    list_clientes=[]
    for c in clientes:
        list_clientes.append({
            "id":c.id,
            "Cliente":c.Cliente
        })

    response={}
    response['dato']=list_clientes
    return HttpResponse(json.dumps(response),content_type="application/json")

@login_required
@permission_required(['Operarios.add_relevamientocab', 'Operarios.view_relevamientocab'], raise_exception=True)
def Relevamiento(request, id_puntoServicio=None):
    primeraVez = 0

    try:
        """
        Obtenemos el punto de servicio
        """
        puntoSer = PuntoServicio.objects.get(Q(pk=id_puntoServicio))
    except PuntoServicio.DoesNotExist:
        raise Http404("Punto de Servicio no existe")

    relevamiento = RelevamientoCab.objects.filter(Q(puntoServicio_id = puntoSer.id)).first()
    cabeceraNueva=False
    if relevamiento == None:
        primeraVez = 1
        relevamiento = RelevamientoCab(puntoServicio=puntoSer,cantAprendices=0,cantidad=0)
        cabeceraNueva=True
        relevamiento.save()

       
    relevamientoDetFormSet =        inlineformset_factory(RelevamientoCab, RelevamientoDet, form=RelevamientoDetForm, extra=1, can_delete=True)
    relevamientoEspFormSet =        inlineformset_factory(RelevamientoCab, RelevamientoEsp, form=RelevamientoEspForm, extra=1, can_delete=True)
    relevamientoCupoHorasFormSet =  inlineformset_factory(RelevamientoCab, RelevamientoCupoHoras, form=RelevamientoCupoHorasForm, extra=1, can_delete=True)
    relevamientoMensuFormSet =      inlineformset_factory(RelevamientoCab, RelevamientoMensualeros, form=RelevamientoMensualerosForm, extra=1, can_delete=True)

    if request.method == 'POST':
        if  (request.POST.get('action') == 'add_det'):
            form = RelevamientoForm(request.POST, instance=relevamiento)
            relevamDetFormSet = relevamientoDetFormSet(request.POST, instance=relevamiento)
            relevamEspFormSet = relevamientoEspFormSet(request.POST, instance=relevamiento)
            relevamCuHrFormSet = relevamientoCupoHorasFormSet(request.POST, instance=relevamiento)
            relevamMenFormSet = relevamientoMensuFormSet(request.POST, instance=relevamiento)
        else:
            form = RelevamientoForm(request.POST, instance=relevamiento)
            relevamDetFormSet = relevamientoDetFormSet(request.POST, instance=relevamiento)
            relevamEspFormSet = relevamientoEspFormSet(request.POST, instance=relevamiento)
            relevamCuHrFormSet = relevamientoCupoHorasFormSet(request.POST, instance=relevamiento)
            relevamMenFormSet = relevamientoMensuFormSet(request.POST, instance=relevamiento)
            i = 0
            post_mutable = request.POST.copy()
            for form in relevamMenFormSet:
                if request.POST.get('relevamientomensualeros_set-'+str(i)+'-sueldo') != 'None':
                    sueldoMask = request.POST.get('relevamientomensualeros_set-'+str(i)+'-sueldo')
                    end = len(sueldoMask)
                    sueldoMask = sueldoMask[3:end]
                    sueldoUnmask = sueldoMask.replace('.','')
                    post_mutable['relevamientomensualeros_set-'+str(i)+'-sueldo'] =sueldoUnmask
                i+=1
            form = RelevamientoForm(post_mutable, instance=relevamiento)
            relevamMenFormSet = relevamientoMensuFormSet(post_mutable, instance=relevamiento)

            print(form.errors)
            print(relevamDetFormSet.errors)
            print(relevamEspFormSet.errors)
            print(relevamCuHrFormSet.errors)
            print(relevamMenFormSet.errors)

            if form.is_valid() and relevamDetFormSet.is_valid() and relevamEspFormSet.is_valid() and relevamCuHrFormSet.is_valid() and relevamMenFormSet.is_valid():
                
                #update de la cabecera, retorna el nuevo ID
                #print(relevamDetFormSet.cleaned_data)
                #Estado y fecha setee en duro
                emptyvar={}
                
                rel_det="[ "
                for item in relevamDetFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ):
                        rel_det+=str({
                                'relevamientocab_id':str(relevamiento.id),
                                'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                                'tipoServPart':str(item.get('tipoServPart')),
                                'orden': str(item.get('orden')),
                                'lunEnt':str(item.get('lunEnt')),
                                'lunSal':str(item.get('lunSal')),
                                'marEnt':str(item.get('marEnt')),
                                'marSal':str(item.get('marSal')),
                                'mieEnt':str(item.get('mieEnt')),
                                'mieSal':str(item.get('mieSal')),
                                'jueEnt':str(item.get('jueEnt')),
                                'jueSal':str(item.get('jueSal')),
                                'vieEnt':str(item.get('vieEnt')),
                                'vieSal':str(item.get('vieSal')),
                                'sabEnt':str(item.get('sabEnt')),
                                'sabSal':str(item.get('sabSal')),
                                'domEnt':str(item.get('domEnt')),
                                'domSal':str(item.get('domSal')),
                                'DELETE':str(item.get('DELETE'))})
                        rel_det+=","
                rel_det=rel_det[:-1]
                rel_det+="]"

                rel_men="[ "
                for item in relevamMenFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ) and not ((item.get('mensuCantidad') is None or item.get('mensuCantidad') <=0 ) and ( item.get('sueldo') is None or item.get('sueldo') <=0) ):
                        rel_men+=str({
                            'relevamientocab_id':str(relevamiento.id),
                            'mensuCantidad':(0 if item.get('mensuCantidad') is None else str(item.get('mensuCantidad'))),
                            'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                            'sueldo':str(item.get('sueldo')),
                            'DELETE':str(item.get('DELETE'))})
                        rel_men+=","
                rel_men=rel_men[:-1]
                rel_men+="]"


                rel_cup="[ "
                for item in relevamCuHrFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ):
                        rel_cup+=str({	
                            'relevamientocab_id':str(relevamiento.id),
                            'cantCHoras':str(item.get('cantCHoras')),
                            'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                            'frecuencia':str(item.get('frecuencia')),
                            'tipoHora':str(item.get('tipoHora')),
                            'DELETE':str(item.get('DELETE'))})
                        rel_cup+=","
                rel_cup=rel_cup[:-1]
                rel_cup+="]"


                rel_esp="[ "
                for item in relevamEspFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ):
                        rel_esp+=str({
                            'relevamientocab_id':str(relevamiento.id),
                            'tipoServicio':str(item.get('tipo')),
                            'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                            'frecuencia':str(item.get('frecuencia')),
                            'cantHoras':str(item.get('cantHoras')),
                            'DELETE':str(item.get('DELETE'))})
                        rel_esp+=","
                rel_esp=rel_esp[:-1]
                rel_esp+="]"

                conn= connection.cursor()
                params=(
                    str({'id': str(relevamiento.id),'puntoServicio_id':str(puntoSer.id),
                        'cantidad': int(0 if form.cleaned_data.get('cantidad') is None else form.cleaned_data.get('cantidad')),
                        'cantAprendices': int(0 if form.cleaned_data.get('cantAprendices') is None else form.cleaned_data.get('cantAprendices')),
                        'cantidadHrTotal': 0 if form.cleaned_data.get('cantidadHrTotal') is None else form.cleaned_data.get('cantidadHrTotal'),
                        'cantidadHrEsp': 0 if form.cleaned_data.get('cantidadHrEsp') is None else form.cleaned_data.get('cantidadHrEsp'),
                        'fechaInicio':str(form.cleaned_data.get('fechaInicio')),
                        'tipoSalario':str(form.cleaned_data.get('tipoSalario')),
                        'comentario': str(form.cleaned_data.get('comentario'))}).replace('\'','\"'),
                    rel_det.replace('\'','\"'),
                    rel_men.replace('\'','\"'),
                    rel_cup.replace('\'','\"'), 
                    rel_esp.replace('\'','\"'),
                    0)
                print(params)
                
                conn.execute('relevamiento_manager %s,%s,%s,%s,%s,%s ',params)
                result = conn.fetchone()[0]
                print(result)
                conn.close()
                
                if result ==0:
                    messages.success(request, 'Servicio aprobado creado correctamente.')
                    return redirect('Operarios:servicio_aprobado')
                else:
                    messages.warning(request, 'No se pudo guardar los cambios')    
            else:
                messages.warning(request, 'No se pudo guardar los cambios')
    else:
        """
        Seteamos el punto de servicio
        """
        relevamiento.puntoServicio = puntoSer

        if primeraVez == 1:
            form = RelevamientoForm(instance=relevamiento, initial={'tipoSalario': 1})
        else:
            form = RelevamientoForm(instance=relevamiento)

        relevamDetFormSet =     relevamientoDetFormSet(instance=relevamiento)
        relevamEspFormSet =     relevamientoEspFormSet(instance=relevamiento)
        relevamCuHrFormSet =    relevamientoCupoHorasFormSet(instance=relevamiento)
        relevamMenFormSet =     relevamientoMensuFormSet(instance=relevamiento)

    contexto = {
            'title': 'Servicio Aprobado',
            'form': form,
            'relevamDetFormSet': relevamDetFormSet,
            'relevamEspFormSet': relevamEspFormSet,
            'relevamCuHrFormSet': relevamCuHrFormSet,
            'relevamMenFormSet': relevamMenFormSet,
        }

    return render(request, 'puntoServicio/puntoServicio_relevamiento.html', context=contexto)


@login_required
@permission_required('Operarios.add_operario', raise_exception=True)
def Operarios_create(request):
    ciudades = Ciudad.objects.all()
    nacionalidades = Nacionalidad.objects.all()
    if request.method == 'POST': 
        form = OperarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Operario creado correctamente.')
        else:
            messages.warning(request, 'No se pudo cargar el Operario')
        return redirect('Operarios:operarios_list')
    else:
        form = OperarioForm()
        contexto = {
            'title': 'Nuevo Operario',
            'form': form,
            'ciudades':ciudades,
            'nacionalidades':nacionalidades
        }
    
    return render(request, 'operarios/operarios_form.html', context=contexto)

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def Operarios_list(request):
    operarios = Operario.objects.all()
    contexto = {'Operarios': operarios}
    return render(request, 'operarios/operarios_list.html', context=contexto)

@login_required
@permission_required('Operarios.change_operario', raise_exception=True)
def Operarios_update(request, pk):
    operario = Operario.objects.get(id=pk)
    print(operario.FechaNacimiento)
    FechaNacimiento = operario.FechaNacimiento.strftime("%d/%m/%Y")
    FechaInicio = operario.FechaInicio.strftime("%d/%m/%Y")
    ciudades = Ciudad.objects.all()
    nacionalidades = Nacionalidad.objects.all()
    if request.method == 'GET':
        form = OperarioForm(instance=operario)
        contexto = {
            'title': 'Editar Operario',
            'form': form,
            'operario':operario ,
            'ciudades':ciudades,
            'FechaNacimiento':FechaNacimiento,
            'FechaInicio':FechaInicio,
            'nacionalidades':nacionalidades
        }
    else:
        form = OperarioForm(request.POST, instance=operario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Operario modificado correctamente.')
        return redirect('Operarios:operarios_list')

    return render(request, 'operarios/operarios_form.html', context=contexto)

@login_required
@permission_required('Operarios.delete_operario', raise_exception=True)
def Operarios_delete(request, pk):
    operarios = Operario.objects.get(id=pk)
    if request.method == 'POST':
        operarios.delete()
        messages.warning(request, 'Operario eliminado correctamente')
        return redirect('Operarios:operarios_list')
    return render(request, 'operarios/operarios_delete.html', {'operarios': operarios})

def getPuntosServicios(request):
    puntoServi = PuntoServicio.objects.all()
    puntos =[]
    i=1
    for p in puntoServi:
        totalHora=""
        horasAsig=""
        horas=""
        minutos=""
        horasRestante=""
        minutosRestante=""
        cantidadMinutos=""
        estado=""
        if RelevamientoCab.objects.filter(Q(puntoServicio_id=p.id) ).exists():
            relevamientoCab = RelevamientoCab.objects.get(Q(puntoServicio_id=p.id))
            totalHora = relevamientoCab.cantidadHrTotal
        if AsignacionCab.objects.filter(Q(puntoServicio_id=p.id) ).exists():
            asignacionCab = AsignacionCab.objects.get(Q(puntoServicio_id=p.id) )
            estado = asignacionCab.reAsignar
            horasAsig = asignacionCab.totalasignado
        if  totalHora and horasAsig:
            horasTotales,minutosTotales = totalHora.split(':')
            horasAsignadas,minutosAsignadas = horasAsig.split(':')
            cantidadMinutos = restarHoras( int(horasTotales),int(horasAsignadas),int(minutosTotales),int(minutosAsignadas))

        puntos.append({
            "id":i,
            "idPunto":p.id,
            "puntservnombre":p.NombrePServicio,
            "horatotal":totalHora,
            "horasasignada":horasAsig,
            "horafaltante":cantidadMinutos,
            "estado":estado
        })
        i=i+1

    response={}
    response['dato']=puntos
    return HttpResponse(json.dumps(response),content_type="application/json")


@login_required
@permission_required('Operarios.add_planificacioncab', raise_exception=True)
def Planificacion_create(request, id_puntoServicio=None):
    logging.getLogger("error_logger").error('Se ingreso en el metodo planificacion_create')
    ''' Obtenemos el punto de servicio, en caso de error se muesta un error 404 '''
    try:
        puntoSer = PuntoServicio.objects.get(Q(pk=id_puntoServicio)  )
    except PuntoServicio.DoesNotExist as err:
        logging.getLogger("error_logger").error('Punto de Servicio no existe: {0}'.format(err))
        raise Http404("Punto de Servicio no existe")

    ''' Obtenemos el relevamiento para mostrar en la pantalla '''
    relevamiento = RelevamientoCab.objects.filter(Q(puntoServicio_id = puntoSer.id) ).first()

    ''' Obtenemos la planificacion en caso de que exista una '''
    planificacion = PlanificacionCab.objects.filter(Q(puntoServicio_id = puntoSer.id) ) .first()

    initial = []
    CantlimpiezaProf = 1
    if planificacion == None:
        planificacion = PlanificacionCab(puntoServicio=puntoSer,cantidad=0)
        planificacion.save()
        '''ULTIMA VERSION DE RELEVAMIENTO ESP'''
        print (relevamiento.relevamientoesp_set.all())
        if relevamiento:
            if len(relevamiento.relevamientoesp_set.all())>0:
                for relevesp in relevamiento.relevamientoesp_set.all():
                    initial.append({'tipo': relevesp.tipo, 
                                    'frecuencia': relevesp.frecuencia,
                                    'cantHoras': relevesp.cantHoras})
                    # initial=[
                    #         {'especialista': 2, 
                    #         'tipo': 2,
                    #         'frecuencia': 'ANUAL',
                    #         'dia': 'SAB',
                    #         'cantHoras': 99}
                    #         ]
                CantlimpiezaProf = len(initial)
           

    planificacionOpeFormSet = inlineformset_factory(PlanificacionCab, PlanificacionOpe, form=PlanificacionOpeForm, extra=1, can_delete=True)
    planificacionEspFormSet = inlineformset_factory(PlanificacionCab, PlanificacionEsp, form=PlanificacionEspForm, extra=CantlimpiezaProf, can_delete=True)

    if request.method == 'POST':

        if  (request.POST.get('action') == 'add_det') or (request.POST.get('action') == 'add_esp'):
            form = PlanificacionForm(request.POST, instance=planificacion)
            planifOpeFormSet = planificacionOpeFormSet(request.POST, instance=planificacion)
            planifEspFormSet = planificacionEspFormSet(request.POST, instance=planificacion)
        else:
            form = PlanificacionForm(request.POST, instance=planificacion)
            planifOpeFormSet = planificacionOpeFormSet(request.POST, instance=planificacion)
            planifEspFormSet = planificacionEspFormSet(request.POST, instance=planificacion)

            if form.is_valid(relevamiento.cantidad, relevamiento.cantidadHrTotal, relevamiento.cantidadHrEsp) and planifOpeFormSet.is_valid() and planifEspFormSet.is_valid():
                #form.save()
                #planifOpeFormSet.save()
                #planifEspFormSet.save()

                emptyvar={}
                pln_ope="[ "
                for item in  planifOpeFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ):
                        pln_ope+=str({
                                'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                                'DELETE':str(item.get('DELETE')),
                                'planificacionCab_id': str(planificacion.id),    
                                'especialista':str(item.get('especialista')),
                                'cantidad':str(item.get('cantidad')),
                                'lun':str(item.get('lun')),
                                'mar':str(item.get('mar')),
                                'mie':str(item.get('mie')),
                                'jue':str(item.get('jue')),
                                'vie':str(item.get('vie')),
                                'sab':str(item.get('sab')),
                                'dom':str(item.get('dom')),
                                'fer':str(item.get('fer')),
                                'ent':str(item.get('ent')),
                                'sal':str(item.get('sal')),
                                'corte':str(item.get('corte')),
                                'total':str(item.get('total'))
                                })
                        pln_ope+=","
                pln_ope=pln_ope[:-1]
                pln_ope+="]"

                pln_esp="[ "
                for item in  planifEspFormSet.cleaned_data:
                    if item != emptyvar and not (item.get('id') is None and item.get('DELETE') is True ):
                        pln_esp+=str({
                                'id':str(str(item.get('id').id) if item.get('id') is not None else 'None'),
                                'DELETE':str(item.get('DELETE')),  
                                'planificacionCab_id': str(planificacion.id),    
                                'especialista':str(item.get('especialista')),
                                'tipo':str(item.get('tipo')),
                                'frecuencia':str(item.get('frecuencia')),
                                'cantHoras':str(item.get('cantHoras')),
                                'fechaLimpProf':str(item.get('fechaLimpProf'))
                                })
                        pln_esp+=","
                pln_esp=pln_esp[:-1]
                pln_esp+="]"
                conn= connection.cursor()
                params=(
                    str({'id': str(planificacion.id),'puntoServicio_id':str(puntoSer.id),
                        	'puntoServicio':str(form.cleaned_data.get('puntoServicio')),
                            'cantidad':str(form.cleaned_data.get('cantidad')),
                            'cantHoras':str(form.cleaned_data.get('cantHoras')),
                            'cantHorasNoc':str(form.cleaned_data.get('cantHorasNoc')),
                            'cantHorasEsp':str(form.cleaned_data.get('cantHorasEsp'))                             
                        }).replace('\'','\"'),
                    pln_ope.replace('\'','\"'),
                    pln_esp.replace('\'','\"'),
                    0)
                print(params)
                conn.execute('planificacion_manager %s,%s,%s,%s ',params)
                result = conn.fetchone()[0]
                conn.close()
                print(result)
                if result==0:
                    messages.success(request, 'Se guardo correctamente la planificación')
                    return redirect('Operarios:planificar_list')
                else:
                    messages.warning(request, 'No se pudo guardar los cambios')  
            else:
                messages.warning(request, 'No se pudo guardar los cambios')
    else:
        """
        Seteamos el punto de servicio
        """
        planificacion.puntoServicio = puntoSer

        form = PlanificacionForm(instance=planificacion)
        planifOpeFormSet = planificacionOpeFormSet(instance=planificacion)
        if len(initial) > 0:
            planifEspFormSet = planificacionEspFormSet(instance=planificacion, initial=initial)
        else:
            planifEspFormSet = planificacionEspFormSet(instance=planificacion)

    contexto = {
            'title': 'Nueva Planificación',
            'form': form,
            'planifOpeFormSet': planifOpeFormSet,
            'planifEspFormSet': planifEspFormSet,
            'relevamiento' : relevamiento,
            'puntoServicio' : puntoSer
        }

    return render(request, 'planificacion/planificacion_crear.html', context=contexto)

@login_required
@permission_required('Operarios.view_planificacioncab', raise_exception=True)
def Planificacion_list(request):
    if request.method == 'POST':
        pk_puntoServSeleccionado = request.POST.get('plani_puntoServ')
        return redirect('Operarios:planificar_create', id_puntoServicio=pk_puntoServSeleccionado)
    else:
        puntoServi = PuntoServicio.objects.all()
        contexto = {'PuntosServicio': puntoServi}
        return render(request, 'planificacion/planificacion_list.html', context=contexto)

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def Jefes_list(request):
    jefes = User.objects.filter(cargoasignado__cargo__cargo='Jefe de Operaciones')
    contexto = {'Jefes': jefes}
    return render(request, 'jefes/jefes_list.html', context=contexto)

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def JefesAsignar_list(request):

    jefes = User.objects.filter(cargoasignado__cargo__cargo='Jefe de Operaciones')
    if request.GET.get('first_name'):
        jefes=jefes.filter(first_name__contains=request.GET.get('first_name'))
    if request.GET.get('last_name'):
        jefes=jefes.filter(last_name__contains=request.GET.get('last_name'))

    paginado=Paginator(jefes.order_by('first_name').values("id", "first_name", "last_name"),  request.GET.get('pageSize'))
    listaPaginada=paginado.page(request.GET.get('pageIndex')).object_list
    dataJefes=list(listaPaginada)

    """
    Filtro nuevo
    """
    lista=dataJefes
    response_data={}
    response_data["data"]=lista
    response_data["itemsCount"]=len(jefes)
    return JsonResponse(response_data)

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def FiscalAsignar_list(request):

    fiscal = User.objects.filter(cargoasignado__cargo__cargo='Fiscal')
    if request.GET.get('first_name'):
        fiscal=fiscal.filter(first_name__contains=request.GET.get('first_name'))
    if request.GET.get('last_name'):
        fiscal=fiscal.filter(last_name__contains=request.GET.get('last_name'))

    # paginado=Paginator(fiscal.order_by('last_name').values("id","first_name", "last_name"),  request.GET.get('pageSize'))
    # listaPaginada=paginado.page(request.GET.get('pageIndex')).object_list
    # dataFiscal=list(listaPaginada)

    """
    Filtro nuevo
    """
   
    return HttpResponse(serializers.serialize("json",fiscal ), content_type = 'application/json', status = 200);


@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def Fiscales_list(request):
    fiscales = User.objects.filter(cargoasignado__cargo__cargo='Fiscal')
    contexto = {'Fiscales':fiscales} 
    return render(request, 'fiscales/fiscales_list.html', context=contexto) 

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def Jefes_asig(request, id_user_jefe=None, id_user_fiscal=None): 

    try:
        user_jefe = User.objects.get(pk=id_user_jefe)   
    except User.DoesNotExist as err:
        logging.getLogger("error_logger").error('Usuario de Jefe de Operaciones no existe: {0}'.format(err))

    fiscales_asig = User.objects.filter(Q(FiscalAsigJefeFiscal__userJefe=id_user_jefe))
    consulta = User.objects.filter(FiscalAsigJefeFiscal__userJefe=id_user_jefe).query
    logging.getLogger("error_logger").error('La consulta ejecutada es: {0}'.format(consulta))

    ''' El if siguiente es utilizado en la version anterior de asignacion de fiscales'''   
    # if id_user_fiscal != None:
    #     print("Hola, aca obtengo los fiscales")
    #     user_fiscal = User.objects.get(pk=id_user_fiscal)
    #     asignacion = AsigJefeFiscal(userJefe=user_jefe, userFiscal=user_fiscal)
    #     asignacion.save()
    
    #se trae los fiscales disponibles
    fiscales_disp = User.objects.filter(FiscalAsigJefeFiscal__userJefe__isnull=True, cargoasignado__cargo__cargo='Fiscal')
    consulta2 = User.objects.filter(FiscalAsigJefeFiscal__userJefe__isnull=True, cargoasignado__cargo__cargo='Fiscal').query
    logging.getLogger("error_logger").error('La consulta de fiscales disponibles ejecutada es: {0}'.format(consulta2))


    logging.getLogger("error_logger").error('La consulta de fiscales disponibles ejecutada es: {0}'.format(consulta2))
    #cargamos el contexto
    contexto = {'Fiscales': fiscales_asig,
                'Jefe': user_jefe,
                'Fiscales_disp': fiscales_disp
            }
    return render(request, 'jefes/jefes_asig.html', context=contexto)

@permission_required('Operarios.view_operario', raise_exception=True)
def asignarFiscales(request,id_user_jefe=None ):

    try:
        user_jefe = User.objects.get(pk=id_user_jefe)   
    except User.DoesNotExist as err:
        logging.getLogger("error_logger").error('Usuario de Jefe de Operaciones no existe: {0}'.format(err))

    #Se traen todos los fiscales que estan asignados al jefe de operaciones en cuestion
    fiscales_asig = User.objects.filter(FiscalAsigJefeFiscal__userJefe=id_user_jefe)
    consulta = User.objects.filter(FiscalAsigJefeFiscal__userJefe=id_user_jefe).query
    logging.getLogger("error_logger").error('La consulta ejecutada es: {0}'.format(consulta))
    
    '''Se obtienen todos los ids seleccionados'''
    if request.method == 'POST':
        if request.POST.getlist('fiscales_disp') != None:
            print(request.POST.getlist('fiscales_disp')) 
            conn= connection.cursor()
            params=(id_user_jefe,0)
            conn.execute('asigjefefisc_des_trigger %s,%s',params)
            result = conn.fetchone()[0]
            conn.close()
            if result==0:
                    funciona=True;
                    for fd_id in request.POST.getlist('fiscales_disp'):
                        conn= connection.cursor()
                        params=(id_user_jefe,fd_id,0,0)
                        conn.execute('asig_jefeyfiscal_trigger %s,%s,%s,%s',params)
                        result = conn.fetchone()[0]
                        conn.close()
                        if result==1:
                            funciona=False;
                            messages.success(request, 'Error al modificar las asignaciones.')                                    
            else:
                messages.success(request, 'Error al modificar las asignaciones.')  

    return redirect('Operarios:jefes_list')

@login_required
@permission_required('Operarios.delete_operario', raise_exception=True)
def Jefes_delete(request, id_user_jefe=None, id_user_fiscal=None):
    user_fiscal = User.objects.get(pk=id_user_fiscal)
    user_jefe = User.objects.get(pk=id_user_jefe)

    if request.method == 'POST':
        asignacion = AsigJefeFiscal.objects.get(userJefe=user_jefe, userFiscal=user_fiscal)
        asignacion.delete()
        messages.warning(request, 'Asignación eliminada correctamente')
        return redirect('Operarios:jefes_asigFiscales', id_user_jefe=id_user_jefe)
    
    #cargamos el contexto
    contexto = {'Jefe': user_jefe,
                'Fiscal': user_fiscal
                }
    return render(request, 'jefes/jefes_delete.html', context=contexto)

@login_required
@permission_required('Operarios.view_operario', raise_exception=True)
def Fiscales_asig(request, id_user_fiscal=None, id_puntoServicio=None):
    try:
        user_fiscal = User.objects.get(pk=id_user_fiscal)   
    except User.DoesNotExist as err:
        logging.getLogger("error_logger").error('Usuario de Fiscal no existe: {0}'.format(err))

    if id_puntoServicio != None:
        puntoServicio = PuntoServicio.objects.get(Q(pk=id_puntoServicio) )
        asignacion = AsigFiscalPuntoServicio(userFiscal=user_fiscal, puntoServicio=puntoServicio)
        asignacion.save()

    #Se traen todos los puntos de servicio que estan asignados al fiscal en cuestion
    puntosServ_asig = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=id_user_fiscal))
    consulta = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=id_user_fiscal) ).query
    logging.getLogger("error_logger").error('La consulta ejecutada es: {0}'.format(consulta))
    print("Asignados",puntosServ_asig)
   
    #se trae los puntos de servicio disponibles 
    puntosServ_disp = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal__isnull=True) )
    consulta2 = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal__isnull=True) ).query
    logging.getLogger("error_logger").error('La consulta de puntos de servicio disponibles ejecutada es: {0}'.format(consulta2))

    #cargamos el contexto
    contexto = {'PuntosSer': puntosServ_asig,
                'Fiscal': user_fiscal,
                'PuntosSer_disp': puntosServ_disp
            }
    return render(request, 'fiscales/fiscales_asig.html', context=contexto)


@permission_required('Operarios.view_operario', raise_exception=True)
def asignarPuntosServicio(request,id_user_fiscal=None):

    try:
        user_fiscal = User.objects.get(pk=id_user_fiscal)   
    except User.DoesNotExist as err:
        logging.getLogger("error_logger").error('Usuario de Fiscal no existe: {0}'.format(err))

    #Se traen todos los puntos de servicio que estan asignados al fiscal en cuestion
    puntosServ_asig = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=id_user_fiscal) )
    consulta = PuntoServicio.objects.filter(Q(puntoServicioAsigFiscalPuntoServicio__userFiscal=id_user_fiscal) ).query
    logging.getLogger("error_logger").error('La consulta ejecutada es: {0}'.format(consulta))
    
    '''Se obtienen todos los ids seleccionados'''
    if request.method == 'POST':
        if request.POST.getlist('puntos_disp') != None:
            conn= connection.cursor()
            params=(id_user_fiscal,0)
            conn.execute('asigpsfisc_des_trigger %s,%s',params)
            result = conn.fetchone()[0]
            conn.close()
            if result==0:
                    funciona=True;
                    for ps_id in request.POST.getlist('puntos_disp'):
                        conn= connection.cursor()
                        params=(id_user_fiscal,ps_id,0,0)
                        conn.execute('asig_psfiscal_trigger %s,%s,%s,%s',params)
                        result = conn.fetchone()[0]
                        conn.close()
                        if result==1:
                            funciona=False;
                            messages.success(request, 'Error al modificar las asignaciones.')                         
            else:
                messages.success(request, 'Error al modificar las asignaciones.') 

    return redirect('Operarios:fiscales_list')

@login_required
@permission_required('Operarios.delete_operario', raise_exception=True)
def Fiscales_delete(request, id_user_fiscal=None, id_puntoServicio=None):
    user_fiscal = User.objects.get(pk=id_user_fiscal) 
    puntoServicio = PuntoServicio.objects.get( Q(pk=id_puntoServicio) )

    if request.method == 'POST':
        asignacion = AsigFiscalPuntoServicio.objects.get(userFiscal=user_fiscal, puntoServicio=puntoServicio)
        asignacion.delete()
        messages.warning(request, 'Asignación eliminada correctamente')
        return redirect('Operarios:fiscales_asig', id_user_fiscal=id_user_fiscal)
    
    #cargamos el contexto
    contexto = {'Fiscal': user_fiscal,
                'PuntoServicio': puntoServicio,
                }
    return render(request, 'fiscales/fiscales_delete.html', context=contexto)

def obtenerMarcacion(request):
    return render(request, 'marcacion/marcacion_list.html');

def obtenerFeriado(request):
    return render(request, 'feriados/feriados_list.html');

def getMarcaciones(request):
        marcacion = EsmeEmMarcaciones.objects.all();
        if request.GET.get('codoperacion')  is not None and request.GET.get('codoperacion')!='':
            marcacion=marcacion.filter(codoperacion__contains = request.GET.get('codoperacion'));
        if request.GET.get('codpersona')  is not None and request.GET.get('codpersona')!='':
            marcacion=marcacion.filter(codpersona=request.GET.get('codpersona'));
        if request.GET.get('codcategoria')  is not None  and request.GET.get('codcategoria')!='':
            marcacion=marcacion.filter(codcategoria=request.GET.get('codcategoria'));
        if request.GET.get('numlinea')  is not None  and request.GET.get('numlinea')!='':
            marcacion=marcacion.filter(numlinea__contains = request.GET.get('numlinea'));
        if request.GET.get('codubicacion')  is not None and request.GET.get('codubicacion')!='':
            marcacion=marcacion.filter(codubicacion__contains = request.GET.get('codubicacion'));
        if request.GET.get('fecha')  is not None and request.GET.get('fecha')!='':
            dtd=datetime.strptime(request.GET.get('fecha'),'%d/%m/%Y');
            dtd=dtd.replace(hour=0,minute=0,second=0,microsecond=0);
            marcacion=marcacion.filter(fecha__gte=dtd);
            dtd=dtd.replace(hour=23,minute=59,second=59);
            marcacion=marcacion.filter(fecha__lte=dtd);
            print(dtd)
        if request.GET.get('estado') is not None and request.GET.get('estado')!='':
            marcacion=marcacion.filter(estado__contains = request.GET.get('estado'));
        """paginado=Paginator(marcacion.order_by('fecha'),  request.GET.get('pageSize'))
        listaPaginada=paginado.page(request.GET.get('pageIndex')).object_list
        dataMarcacion=list(listaPaginada)"""
        
        """lista=serializers.serialize("json",dataMarcacion )"""
       

        """
        Filtro nuevo
        """
        """
        response_data={}
        response_data["data"]=json.loads(serializers.serialize("json",dataMarcacion )),
        response_data["itemsCount"]=len(marcacion)
        
        print (response_data)
        print ("Serializado")
        
       
        
        return HttpResponse(json.dumps(response_data), content_type = 'application/json', status = 200);
        """
        return HttpResponse(serializers.serialize("json",marcacion ), content_type = 'application/json', status = 200);


def default(o):
    if isinstance(o,(datetime.date,datetime.datetime)):
        return o.isoformat();
    if isinstance(o,decimal.Decimal):
        return str(o);

def getMarcacionesPaginada(request):
        marcacion = EsmeEmMarcaciones.objects.all();
        if request.GET.get('codoperacion')  is not None and request.GET.get('codoperacion')!='':
            marcacion=marcacion.filter(codoperacion__contains = request.GET.get('codoperacion'));
        if request.GET.get('codpersona')  is not None and request.GET.get('codpersona')!='':
            marcacion=marcacion.filter(codpersona=request.GET.get('codpersona'));
        if request.GET.get('codcategoria')  is not None  and request.GET.get('codcategoria')!='':
            marcacion=marcacion.filter(codcategoria=request.GET.get('codcategoria'));
        if request.GET.get('numlinea')  is not None  and request.GET.get('numlinea')!='':
            marcacion=marcacion.filter(numlinea__contains = request.GET.get('numlinea'));
        if request.GET.get('codubicacion')  is not None and request.GET.get('codubicacion')!='':
            marcacion=marcacion.filter(codubicacion__contains = request.GET.get('codubicacion'));
        if request.GET.get('fecha')  is not None and request.GET.get('fecha')!='':
            dtd=datetime.strptime(request.GET.get('fecha'),'%d/%m/%Y');
            dtd=dtd.replace(hour=0,minute=0,second=0,microsecond=0);
            marcacion=marcacion.filter(fecha__gte=dtd);
            dtd=dtd.replace(hour=23,minute=59,second=59);
            marcacion=marcacion.filter(fecha__lte=dtd);
            print(dtd)
        if request.GET.get('estado') is not None and request.GET.get('estado')!='':
            marcacion=marcacion.filter(estado__contains = request.GET.get('estado'));
        
        pagina=1;
        items=10;
        cuenta=marcacion.count();

        if request.GET.get('pagina') is not None and request.GET.get('pagina')!='':
            pagina=int(request.GET.get('pagina'));
        if request.GET.get('items') is not None and request.GET.get('items')!='':
            items=int(request.GET.get('items'));
        marcacion=marcacion[(pagina-1)*items:pagina*items];
        marcacion=[model_to_dict(item) for item in marcacion];
        respuesta={};
        respuesta['lista']=marcacion;
        respuesta['totalPaginas']=round(cuenta/items);
        respuesta['paginaActual']=pagina;
        respuesta['totalItems']=cuenta;
        respuesta['items']=items;

        return HttpResponse(json.dumps(respuesta,default=default), content_type = 'application/json', status = 200);
        


def descargarMarcaciones(request):
        marcacion = EsmeEmMarcaciones.objects.all();
        if request.GET.get('codoperacion')  is not None and request.GET.get('codoperacion')!='':
            marcacion=marcacion.filter(codoperacion__contains = request.GET.get('codoperacion'));
        if request.GET.get('codpersona')  is not None and request.GET.get('codpersona')!='':
            marcacion=marcacion.filter(codpersona=request.GET.get('codpersona'));
        if request.GET.get('codcategoria')  is not None  and request.GET.get('codcategoria')!='':
            marcacion=marcacion.filter(codcategoria=request.GET.get('codcategoria'));
        if request.GET.get('numlinea')  is not None  and request.GET.get('numlinea')!='':
            marcacion=marcacion.filter(numlinea__contains = request.GET.get('numlinea'));
        if request.GET.get('codubicacion')  is not None and request.GET.get('codubicacion')!='':
            marcacion=marcacion.filter(codubicacion__contains = request.GET.get('codubicacion'));
        if request.GET.get('fecha')  is not None and request.GET.get('fecha')!='':
            dtd=datetime.strptime(request.GET.get('fecha'),'%Y-%m-%dT%H:%M:%S.%fZ');
            dtd=dtd.replace(hour=0,minute=0,second=0,microsecond=0);
            marcacion=marcacion.filter(fecha__gte=dtd);
            dtd=dtd.replace(hour=23,minute=59,second=59);
            marcacion=marcacion.filter(fecha__lte=dtd);
        if request.GET.get('estado') is not None and request.GET.get('estado')!='':
            marcacion=marcacion.filter(estado__contains = request.GET.get('estado'));

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-marcaciones.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Marcaciones'
        # Define the titles for columns
        columns = [
            'Operacion',
            'Cedula',
            'Codigo Categoria',
            'Numero de Linea',
            'Ubicacion',
            'Fecha',
            'Estado',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for marc in marcacion:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                marc.codoperacion,
                marc.codpersona,
                marc.codcategoria,
                marc.numlinea,
                marc.codubicacion,
                marc.fecha,
                marc.estado
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response;



def getFeriados(request):
        feriados = Feriados.objects.all();
        if request.GET.get('anho')  is not None  and request.GET.get('anho')!='':
            feriados=feriados.filter(anho = request.GET.get('anho'));
        if request.GET.get('descripcion')  is not None and request.GET.get('descripcion')!='':
            feriados=feriados.filter(descripcion__contains = request.GET.get('descripcion'));
        if request.GET.get('fecha')  is not None and request.GET.get('fecha')!='':
            dtd=datetime.strptime(request.GET.get('fecha'),'%d/%m/%Y');
            dtd=dtd.replace(hour=0,minute=0,second=0,microsecond=0);
            feriados=feriados.filter(fecha__gte=dtd);
            dtd=dtd.replace(hour=23,minute=59,second=59);
            print(dtd)
            feriados=feriados.filter(fecha__lte=dtd);
        return HttpResponse(serializers.serialize('json', feriados), content_type = 'application/json', status = 200);
def makeFeriados(request):
        print(request.POST)
        response={}
        try:
            Feriados.objects.create(
                anho= request.POST.get("anho"),
                fecha=datetime.strptime(request.POST.get('fecha'),'%Y-%m-%d'),
                descripcion= request.POST.get("descripcion"))
            response['dato']=[]
            response['codigo']=0
            response['mensaje']="Se creó el registro"
            return HttpResponse(json.dumps(response),content_type="application/json")
        except Exception as err:
            logging.getLogger("error_logger").error('No se pudo crear el registro: {0}'.format(err))
            response['codigo']=1
            response['dato']=[]
            response['mensaje']="No se pudo crear el registro"
            return HttpResponse(json.dumps(response),content_type="application/json")
def editFeriados(request,feriado_id):
        feriado=Feriados.objects.get(id = feriado_id);
        if request.POST.get('anho') is not None:
            feriado.anho=request.POST.get('anho');
        if request.POST.get('descripcion') is not None:
            feriado.descripcion=request.POST.get('descripcion');
        if request.POST.get('fecha') is not None:
            feriado.fecha=datetime.strptime(request.POST.get('fecha'),'%Y-%m-%dT%H:%M:%S.%fZ');
        feriado.save();
        return HttpResponse(status = 201)
def deleteFeriados(request,feriado_id):
        feriado=Feriados.objects.get(id = feriado_id);
        feriado.delete();
        return HttpResponse(status = 201)


class EsmeEmMarcacionesClass(ListView):

    def index(request):
        return render(request, 'marcacion/marcacion_list.html');

    

    def post(self, request):
        EsmeEmMarcaciones.objects.create(
            codoperacion= request.POST.get("codoperacion"),
            codpersona= request.POST.get("codpersona"),
            codcategoria= request.POST.get("codcategoria"),
            numlinea= request.POST.get("numlinea"),
            codubicacion= request.POST.get("codubicacion"),
            fecha= request.POST.get("fecha"),
            estado= request.POST.get("estado")

        )
        return HttpResponse(status = 201)

    def put(self, request, marcacion_id):
        marcacion = EsmeEmMarcaciones.objects.get(idpersonaevento = marcacion_id)
        marcacion.codoperacion= request.PUT.get("codoperacion"),
        marcacion.codpersona= request.PUT.get("codpersona"),
        marcacion.codcategoria= request.PUT.get("codcategoria"),
        marcacion.numlinea= request.PUT.get("numlinea"),
        marcacion.codubicacion= request.PUT.get("codubicacion"),
        marcacion.fecha= request.PUT.get("fecha"),
        marcacion.estado= request.PUT.get("estado")
        marcacion.save()
        return HttpResponse(status = 200)

    def delete(self, request, marcacion_id):
        marcacion = EsmeEmMarcaciones.objects.get(idpersonaevento = marcacion_id)
        marcacion.delete()
        return HttpResponse(status = 200)

    def to_json(self, objects):
        return serializers.serialize('json', objects)
def restarHoras(totalHora,asigHora,totalMin,asigMin):
    totalHorasMinutos = totalHora*60
    totalAsigHorasMinutos = asigHora*60
    cantidadTotalDeMinutos = (totalHorasMinutos+totalMin)-(totalAsigHorasMinutos+asigMin)
    cantidadTotalHoras = cantidadTotalDeMinutos//60
    cantidadTotalDeMinutos = cantidadTotalDeMinutos%60
    return "{}:{}".format(cantidadTotalHoras,int(cantidadTotalDeMinutos))

def vicios(request):
    puntoServi = PuntoServicio.objects.all()
    puntos =[]
    i=1
    for p in puntoServi:
        totalHora=""
        horasAsig=""
        horas=""
        minutos=""
        horasRestante=""
        minutosRestante=""
        cantidadMinutos=""
        estado=False
        if RelevamientoCab.objects.filter(Q(puntoServicio_id=p.id) ).exists():
            relevamientoCab = RelevamientoCab.objects.get(Q(puntoServicio_id=p.id) )
            totalHora = relevamientoCab.cantidadHrTotal
        if PlanificacionCab.objects.filter(Q(puntoServicio_id=p.id) ).exists():
            asignacionCab = PlanificacionCab.objects.get(Q(puntoServicio_id=p.id) )
            estado = asignacionCab.rePlanificar
            horasAsig = asignacionCab.cantHoras
        if  totalHora and horasAsig:
            horasTotales,minutosTotales = totalHora.split(':')
            horasAsignadas,minutosAsignadas = horasAsig.split(':')
            cantidadMinutos = restarHoras( int(horasTotales),int(horasAsignadas),int(minutosTotales),int(minutosAsignadas))

        puntos.append({
            "id":i,
            "idPunto":p.id,
            "puntservnombre":p.NombrePServicio,
            "horatotal":totalHora,
            "horasasignada":horasAsig,
            "horafaltante":cantidadMinutos,
            "estado":estado
        })
        i=i+1

    response={}
    response['dato']=puntos
    return HttpResponse(json.dumps(response),content_type="application/json")
