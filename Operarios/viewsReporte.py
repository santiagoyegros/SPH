import logging
from django.shortcuts import render, redirect, render_to_response
from django.core import serializers
from datetime import datetime
from django.http import HttpResponse, Http404, JsonResponse
from django.contrib import messages
from django.urls import reverse_lazy
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required, permission_required
from Operarios.models import Operario, DiaLibre, AsignacionDet, AsignacionCab
from django.db.models import Q
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm,inch,mm
from reportlab.lib.colors import HexColor
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.utils import ImageReader
import json
from datetime import datetime
from PIL import Image
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import user_passes_test
from django.contrib.staticfiles.templatetags.staticfiles import static


decorator_with_arguments = lambda decorator: lambda *args, **kwargs: lambda func: decorator(func, *args, **kwargs)

@decorator_with_arguments
def permiso_requerido(function, perm):
    def _function(request, *args, **kwargs):
        if request.user.has_perm(perm):
            return function(request, *args, **kwargs)
        else:
            return render(request,'forbidden.html')
    return _function

@login_required
@permiso_requerido('Operarios.view_AsigsPorOperario')
def AsigPorOperario(request):
    operarios=Operario.objects.all()
    contexto={
        'operarios':operarios
    }
    return render(request,'reportes/reporte_operarios_asig.html',contexto)


def datosOperario(request):
    operario=Operario.objects.get(Q(id=request.GET.get("id")))
    asignaciones=AsignacionDet.objects.filter(operario_id=operario.id)
    asigAux={}
    listaAsignaciones=[]
    total=0
    diasLibres=getDiaLibre(request.GET.get("id"))
    for a in asignaciones:
        asigAux['id']=int(a.id)
        asigAux['totalHoras']=int(a.totalHoras)
        cab=AsignacionCab.objects.get(id=a.asignacionCab_id)
        asigAux['puntoServicio']=cab.puntoServicio.NombrePServicio
        if(a.lunEnt!=None and a.lunSal!=None):
            h=str(a.lunEnt).split(':')
            asigAux['lunEnt']=h[0]+':'+h[1]
            h=str(a.lunSal).split(':')
            asigAux['lunSal']=h[0]+':'+h[1]
        if(a.marEnt!=None and a.marSal!=None):
            h=str(a.marEnt).split(':')
            asigAux['marEnt']=h[0]+':'+h[1]
            h=str(a.marSal).split(':')
            asigAux['marSal']=h[0]+':'+h[1]
        if(a.mieEnt!=None and a.mieSal!=None):
            h=str(a.mieEnt).split(':')
            asigAux['mieEnt']=h[0]+':'+h[1]
            h=str(a.mieSal).split(':')
            asigAux['mieSal']=h[0]+':'+h[1]
        if(a.jueEnt!=None and a.jueSal!=None):
            h=str(a.jueEnt).split(':')
            asigAux['jueEnt']=h[0]+':'+h[1]
            h=str(a.jueSal).split(':')
            asigAux['jueSal']=h[0]+':'+h[1]
        if(a.vieEnt!=None and a.vieSal!=None):
            h=str(a.vieEnt).split(':')
            asigAux['vieEnt']=h[0]+':'+h[1]
            h=str(a.vieSal).split(':')
            asigAux['vieSal']=h[0]+':'+h[1]
        if(a.sabEnt!=None and a.sabSal!=None):
            h=str(a.sabEnt).split(':')
            asigAux['sabEnt']=h[0]+':'+h[1]
            h=str(a.sabSal).split(':')
            asigAux['sabSal']=h[0]+':'+h[1]
        if(a.domEnt!=None and a.domSal!=None):
            h=str(a.domEnt).split(':')
            asigAux['domEnt']=h[0]+':'+h[1]
            h=str(a.domSal).split(':')
            asigAux['domSal']=h[0]+':'+h[1]
        listaAsignaciones.append(asigAux)
        asigAux=asigAux.copy()
        total=total+int(a.totalHoras)



    i=1
    data=[]
    for a in asignaciones:
        listaAsigs=[]
        cab=AsignacionCab.objects.get(id=a.asignacionCab_id)
        puntoServ=cab.puntoServicio.NombrePServicio
        if(a.lunEnt and a.lunSal):
            h1=str(a.lunEnt).split(':')
            h2=str(a.lunSal).split(':')
            lunes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            lunes=""
        if(a.marEnt and a.marSal):
            h1=str(a.marEnt).split(':')
            h2=str(a.marSal).split(':')
            martes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            martes=""
        if(a.mieEnt and a.mieSal):
            h1=str(a.mieEnt).split(':')
            h2=str(a.mieSal).split(':')
            miercoles=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            miercoles=""
        if(a.jueEnt and a.jueSal):
            h1=str(a.jueEnt).split(':')
            h2=str(a.jueSal).split(':')
            jueves=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            jueves=""
        if(a.vieEnt and a.vieSal):
            h1=str(a.vieEnt).split(':')
            h2=str(a.vieSal).split(':')
            viernes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            viernes=""
        if(a.sabEnt and a.sabSal):
            h1=str(a.sabEnt).split(':')
            h2=str(a.sabSal).split(':')
            sabado=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            sabado=""
        if(a.domEnt and a.domSal):
            h1=str(a.domEnt).split(':')
            h2=str(a.domSal).split(':')
            domingo=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
        else:
            domingo=""
        listaAsigs=[i,puntoServ,a.totalHoras,lunes,martes,miercoles,jueves,viernes,sabado,domingo]
        data.append(listaAsigs)
        i+=1

    diaLibre1=diasLibres[0]["nombre"]
    hora=diasLibres[0]["hEntrada"].split(':')
    hora1=hora[0]+':'+hora[1]
    diaLibre2=diasLibres[len(diasLibres)-1]["nombre"]
    hora=diasLibres[len(diasLibres)-1]["hSalida"].split(':')
    hora2=hora[0]+':'+hora[1]
    data={
        'nombre':operario.nombre,
        'apellido':operario.apellido,
        'legajo':operario.nroLegajo,
        'total':total,
        'diaLibre1':diaLibre1,
        'hora1':hora1,
        'diaLibre2':diaLibre2,
        'hora2':hora2,
        'asignaciones':data
    }
    return HttpResponse(json.dumps(data),content_type="application/json")




@login_required
@permiso_requerido('Operarios.view_AsigsPorOperarioExcel')
def getAsignacionesExcel(request,id_operario=None):
    if(id_operario!='0'):
        operario=Operario.objects.get(Q(id=id_operario))
        asignaciones=AsignacionDet.objects.filter(operario_id=operario.id)
        asigAux={}
        listaAsignaciones=[]
        total=0
        diasLibres=getDiaLibre(id_operario)
        cantAsignaciones=0
        for a in asignaciones:
            cantAsignaciones+=1
            total=total+int(a.totalHoras)


        diaLibre1=diasLibres[0]["nombre"]
        hora=diasLibres[0]["hEntrada"].split(':')
        hora1=hora[0]+':'+hora[1]
        diaLibre2=diasLibres[len(diasLibres)-1]["nombre"]
        hora=diasLibres[len(diasLibres)-1]["hSalida"].split(':')
        hora2=hora[0]+':'+hora[1]


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={operario}-asignaciones.xlsx'.format(
            operario=str(operario.nombre)+" "+str(operario.apellido) + " - "+str(operario.nroLegajo)
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Asignaciones operario - ' + operario.nombre + ' '+operario.apellido+' - '+operario.nroLegajo
        worksheet.row_dimensions[1].height=20
        worksheet.column_dimensions['A'].width=25
        worksheet.column_dimensions['B'].width=20
        worksheet.column_dimensions['C'].width=20
        worksheet.column_dimensions['D'].width=20
        worksheet.column_dimensions['E'].width=25
        worksheet.column_dimensions['F'].width=20
        worksheet.column_dimensions['G'].width=20
        worksheet.column_dimensions['H'].width=20
        worksheet.column_dimensions['I'].width=20
        worksheet.column_dimensions['J'].width=20
        color=PatternFill(start_color='86273e',end_color='86273e',fill_type='solid')

        columns = [
        'Reporte de operario'
        ]
        row_num = 1
        worksheet.row_dimensions[row_num].font=Font(bold=True,color='86273e',size=22)
        worksheet.merge_cells('A1:B2')
        
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        row_num = 2

        worksheet.merge_cells('F1:G3')
        png = r'.\Operarios\static\logo.png'
        img = Image.open(png)
        img = img.resize((220,65),Image.NEAREST)
        img.save(png)        
        img = openpyxl.drawing.image.Image(png)
        worksheet.add_image(img, 'F1')
        worksheet.merge_cells('F4:G4')
        worksheet["F4"]=datetime.today().strftime('%d/%m/%Y')
        worksheet["F4"].font=Font(bold=True)
        worksheet["F4"].alignment=Alignment(horizontal='center')
        
        columns = [
            'Nombre',
            'Apellido',
            'Legajo'
        ]
        row_num += 1

        worksheet.row_dimensions[row_num].font=Font(bold=True,color='86273e')
        worksheet.row_dimensions[row_num].alignment=Alignment(horizontal='center')
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        row = [
            operario.nombre,
            operario.apellido,
            operario.nroLegajo,
        ]
        row_num += 1
        worksheet.row_dimensions[row_num].alignment=Alignment(horizontal='center')
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


        columns = [
        'Día libre inicio',
        'hora',
        'Día libre fin',
        'hora'
        ]
        row_num += 1
        worksheet.row_dimensions[row_num].font=Font(bold=True,color='86273e')
        worksheet.row_dimensions[row_num].alignment=Alignment(horizontal='center')
        
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        
        row = [
            diaLibre1,
            hora1,
            diaLibre2,
            hora2
        ]
        row_num += 1
        worksheet.row_dimensions[row_num].alignment=Alignment(horizontal='center')
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        
        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

            
        for j in range(3,7):
            for i in range(1,5):
                worksheet.cell(row=j, column=i).border = thin_border
                worksheet.cell(row=j, column=i).alignment=Alignment(horizontal='center')
                if(j==3 or j==5):
                    worksheet.cell(row=j, column=i).font=Font(bold=True,color='86273e')
        
        worksheet.cell(row=3, column=4).border = None
        worksheet.cell(row=4, column=4).border = None


        columns = []
        row_num += 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        
        columns = []
        row_num += 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        total=str(total)+" hrs"
        columns = [
            'Total horas asignadas:',
            total
        ]
        row_num += 1
        worksheet["A9"].font=Font(bold=True,size=12,color='86273e')
        worksheet["B9"].font=Font(bold=True,size=12,color='000000')
        worksheet["B9"].alignment=Alignment(horizontal='left')

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

        

        #Titulos para la tabla
        columns = [
            'Nro',
            'Punto de servicio',
            'Total Horas',
            'Lunes',
            'Martes',
            'Miércoles',
            'Jueves',
            'Viernes',
            'Sábado',
            'Domingo'
        ]
        row_num += 1
        for i in range(1,len(columns)+1):
            worksheet.cell(row=row_num, column=i).border = thin_border
            worksheet.cell(row=row_num, column=i).alignment=Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=i).font=Font(bold=True,color='86273e') 
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        i=1
        for a in asignaciones:
            cab=AsignacionCab.objects.get(id=a.asignacionCab_id)
            puntoServ=cab.puntoServicio.NombrePServicio
            if(a.lunEnt and a.lunSal):
                lunes=str(a.lunEnt)+" a "+str(a.lunSal)
            else:
                lunes=""
            if(a.marEnt and a.marSal):
                martes=str(a.marEnt)+" a "+str(a.marSal)
            else:
                martes=""
            if(a.mieEnt and a.mieSal):
                miercoles=str(a.mieEnt)+" a "+str(a.mieSal)
            else:
                miercoles=""
            if(a.jueEnt and a.jueSal):
                jueves=str(a.jueEnt)+" a "+str(a.jueSal)
            else:
                jueves=""
            if(a.vieEnt and a.vieSal):
                viernes=str(a.vieEnt)+" a "+str(a.vieSal)
            else:
                viernes=""
            if(a.sabEnt and a.sabSal):
                sabado=str(a.sabEnt)+" a "+str(a.sabSal)
            else:
                sabado=""
            if(a.domEnt and a.domSal):
                domingo=str(a.domEnt)+" a "+str(a.domSal)
            else:
                domingo=""
            row = [
                i,
                puntoServ,
                a.totalHoras,
                lunes,
                martes,
                miercoles,
                jueves,
                viernes,
                sabado,
                domingo
            ]
            row_num += 1
            

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
            i=i+1
            for j in range(1,len(columns)+1):
                worksheet.cell(row=row_num, column=j).border = thin_border
                worksheet.cell(row=row_num, column=j).alignment=Alignment(horizontal='center')    

        workbook.save(response)
        return response



@login_required
@permiso_requerido('Operarios.view_AsigsPorOperarioPdf')
def getAsignacionesPDF(request,id_operario=None):
    operario=Operario.objects.get(Q(id=id_operario))
    asignaciones=AsignacionDet.objects.filter(operario_id=operario.id)
    asigAux={}
    listaAsignaciones=[]
    total=0
    diasLibres=diasLibres=getDiaLibre(id_operario)
    cantAsignaciones=0
    for a in asignaciones:
        cantAsignaciones+=1
        total=total+int(a.totalHoras)

    diaLibre1=diasLibres[0]["nombre"]
    hora=diasLibres[0]["hEntrada"].split(':')
    hora1=hora[0]+':'+hora[1]
    diaLibre2=diasLibres[len(diasLibres)-1]["nombre"]
    hora=diasLibres[len(diasLibres)-1]["hSalida"].split(':')
    hora2=hora[0]+':'+hora[1]


    
    response=HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename='+operario.nombre+' '+operario.apellido+' - '+operario.nroLegajo+'.pdf'
    buffer = BytesIO()
    c=canvas.Canvas(buffer,pagesize=A4)

    c.setFont('Helvetica-Bold',18)
    c.setFillColor(HexColor('#86273e'))
    c.drawString(30,750,'Reporte de Operario')
    c.setFont('Helvetica',12)
    c.drawString(50,700,'Nombre')
    c.drawString(200,700,'Apellido')
    c.drawString(350,700,'Legajo')
    c.drawString(50,650,'Dia Libre Inicio')
    c.drawString(200,650,'Hora')
    c.drawString(350,650,'Dia Libre Fin')
    c.drawString(470,650,'Hora')
    c.setFillColor(HexColor('#000000'))
    c.drawString(50,680,operario.nombre)
    c.drawString(200,680,operario.apellido)
    c.drawString(350,680,operario.nroLegajo)
    c.drawString(50,630,diaLibre1)
    c.drawString(200,630,hora1)
    c.drawString(350,630,diaLibre2)
    c.drawString(470,630,hora2)
    c.setFont('Helvetica-Bold',14)
    c.setFillColor(HexColor('#86273e'))
    c.drawString(30,580,'Total Horas Asignadas:')
    c.setFillColor(HexColor('#000000'))
    totalAux=str(total)+" hrs."
    c.drawString(200,580,totalAux) 
    png = r'.\static\logo1.png'
    logo=ImageReader(png)
    c.drawImage(logo,480,770,width=90,height=60,mask='auto')
    c.line(30,720,580,720)
    c.line(30,720,30,620)    
    c.line(30,620,580,620)
    c.line(580,720,580,620)   

    date=datetime.today().strftime('%d/%m/%Y')
    c.setFont('Helvetica',14)
    c.drawString(490,750,date)

    page=c.getPageNumber()
    text="%s" %page
    c.drawRightString(100*mm,20*mm,text)

    width,height=A4

    styles=getSampleStyleSheet()
    styleBH=styles["Normal"]
    styleBH.alignment=TA_CENTER
    styleBH.fontSize=10
    styleBH.textColor=HexColor('#86273e')  

    nro=Paragraph('''Nro''',styleBH)
    pSer=Paragraph('''Punto Servicio''',styleBH)
    totalHrs=Paragraph('''Total Horas''',styleBH)
    lns=Paragraph('''Lunes''',styleBH)
    mrts=Paragraph('''Martes''',styleBH)
    mrcls=Paragraph('''Miércoles''',styleBH)
    jvs=Paragraph('''Jueves''',styleBH)
    vrns=Paragraph('''Viernes''',styleBH)
    sbd=Paragraph('''Sábado''',styleBH)
    dmg=Paragraph('''Domingo''',styleBH)

    data=[]
    encabezados=[nro,pSer,totalHrs,lns,mrts,mrcls,jvs,vrns,sbd,dmg]
    data.append(encabezados)

    

    styleN=styles["BodyText"]
    styleN.alignment=TA_CENTER
    styleN.fontSize=8

    i=1
    l=[]
    high=510
    
    for a in asignaciones:
        listaAsigs=[]
        cab=AsignacionCab.objects.get(id=a.asignacionCab_id)
        puntoServ=cab.puntoServicio.NombrePServicio
        if(a.lunEnt and a.lunSal):
            h1=str(a.lunEnt).split(':')
            h2=str(a.lunSal).split(':')
            lunes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            lunes=Paragraph(lunes,styleN)
        else:
            lunes=""
        if(a.marEnt and a.marSal):
            h1=str(a.marEnt).split(':')
            h2=str(a.marSal).split(':')
            martes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            martes=Paragraph(martes,styleN)
        else:
            martes=""
        if(a.mieEnt and a.mieSal):
            h1=str(a.mieEnt).split(':')
            h2=str(a.mieSal).split(':')
            miercoles=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            miercoles=Paragraph(miercoles,styleN)
        else:
            miercoles=""
        if(a.jueEnt and a.jueSal):
            h1=str(a.jueEnt).split(':')
            h2=str(a.jueSal).split(':')
            jueves=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            jueves=Paragraph(jueves,styleN)
        else:
            jueves=""
        if(a.vieEnt and a.vieSal):
            h1=str(a.vieEnt).split(':')
            h2=str(a.vieSal).split(':')
            viernes=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            viernes=Paragraph(viernes,styleN)
        else:
            viernes=""
        if(a.sabEnt and a.sabSal):
            h1=str(a.sabEnt).split(':')
            h2=str(a.sabSal).split(':')
            sabado=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            sabado=Paragraph(sabado,styleN)
        else:
            sabado=""
        if(a.domEnt and a.domSal):
            h1=str(a.domEnt).split(':')
            h2=str(a.domSal).split(':')
            domingo=h1[0]+':'+h1[1]+" a "+h2[0]+':'+h2[1]
            domingo=Paragraph(domingo,styleN)
        else:
            domingo=""
        nro=Paragraph('''Nro''',styleBH)
        puntoServ=Paragraph(puntoServ,styleN)
        totalHrs=Paragraph(a.totalHoras,styleN)
        listaAsigs=[i,puntoServ,totalHrs,lunes,martes,miercoles,jueves,viernes,sabado,domingo]
        data.append(listaAsigs)
        i+=1
        high-=18

    table=Table(data,colWidths=[1 * cm, 2.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
    table.setStyle(TableStyle([
        ('BOX',(0,0),(-1,-1),0.25,colors.black),
        ('GRID',(0,0),(-1,-1),0.3*mm,(0,0,0)),
    ]))

    if(total>0):
        table.wrapOn(c,width,height)
        table.drawOn(c,30,high)

        c.showPage()


    c.save()
    pdf=buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response




def getDiaLibre(id_operario):
    diasLibres=[]
    dia={
    'nombre':'',
    'hEntrada':'',
    'hSalida':''
    }
    
    if DiaLibre.objects.filter(Q(id_operario_id=id_operario)).exists():
        diaLibre=DiaLibre.objects.get(Q(id_operario_id=id_operario))
        if(diaLibre.lunEnt!=None and diaLibre.lunSal!=None):
            dia=dia.copy()
            dia["nombre"]="Lunes"
            dia["hEntrada"]=str(diaLibre.lunEnt)
            dia["hSalida"]=str(diaLibre.lunSal)
            diasLibres.append(dia)
        if(diaLibre.marEnt!=None and diaLibre.marSal!=None):
            dia["nombre"]="Martes"
            dia["hEntrada"]=str(diaLibre.marEnt)
            dia["hSalida"]=str(diaLibre.marSal)
            diasLibres.append(dia)
        if(diaLibre.mieEnt!=None and diaLibre.mieSal!=None):
            dia["nombre"]="Miercoles"
            dia["hEntrada"]=str(diaLibre.mieEnt)
            dia["hSalida"]=str(diaLibre.mieSal)
            diasLibres.append(dia)
        if(diaLibre.jueEnt!=None and diaLibre.jueSal!=None):
            dia=dia.copy()
            dia["nombre"]="Jueves"
            dia["hEntrada"]=str(diaLibre.jueEnt)
            dia["hSalida"]=str(diaLibre.jueSal)
            diasLibres.append(dia)
        if(diaLibre.vieEnt!=None and diaLibre.vieSal!=None):
            dia=dia.copy()
            dia["nombre"]="Viernes"
            dia["hEntrada"]=str(diaLibre.vieEnt)
            dia["hSalida"]=str(diaLibre.vieSal)
            diasLibres.append(dia)
        if(diaLibre.sabEnt!=None and diaLibre.sabSal!=None):
            dia=dia.copy()
            dia["nombre"]="Sábado"
            dia["hEntrada"]=str(diaLibre.sabEnt)
            dia["hSalida"]=str(diaLibre.sabSal)
            diasLibres.append(dia)
        if(diaLibre.domEnt!=None and diaLibre.domSal!=None):
            dia=dia.copy()
            dia["nombre"]="Domingo"
            dia["hEntrada"]=str(diaLibre.domEnt)
            dia["hSalida"]=str(diaLibre.domSal)
            diasLibres.append(dia)


    return diasLibres